import openpyxl # biblioteca openpyxl para manipulação de arquivos Excel
from openpyxl import Workbook # classe Workbook para criar uma nova planilha
import selenium # Selenium para automação de navegador
# driver do Chrome para controle do navegador
from selenium.webdriver import Chrome
# Importa as teclas do teclado para interagir com elementos
from selenium.webdriver.common.keys import Keys
# Importa a função load_workbook para carregar um arquivo Excel existente
from openpyxl import load_workbook
import pyautogui as p # biblioteca PyAutoGUI para automação de interface gráfica
wb = Workbook() # Cria um novo arquivo Excel
nome_arquivo = "meu_arquivo_teste.xlsx" # Define o nome do arquivo a ser salvo
wsl = wb.active # Obtém a planilha ativa dentro do arquivo Excel
wsl.title = 'Planilha Aula' # definindo o título da planilha do Excel
# Salva o arquivo Excel no diretório especificado
wb.save(filename=r'C:\Users\User\Desktop\meu_arquivo_teste.xlsx')
# Cole o endereco da sua area de trabalho aqui acima