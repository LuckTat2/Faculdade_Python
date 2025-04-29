import rpa as r
import pyautogui as p
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import datetime as dt

# === ETAPA 1: Criar nova planilha Excel ===
wb = Workbook()
ws = wb.active
ws.title = 'Planilha Aula'
nome_arquivo = r'C:\Users\User\Desktop\meu_arquivo_teste.xlsx'  # Caminho completo

# === ETAPA 2: Coletar cotação do dólar ===
try:
    r.init(visual_automation=True, chrome_browser=True)
    r.url('https://br.investing.com/currencies/usd-brl')
    time.sleep(5)  # Aguarda a página carregar

    # XPath atualizado (confira se está certo no dia da execução)
    xpath_dolar = '//*[@data-test="instrument-price-last"]'
    dolar = r.read(xpath_dolar)
    print(f"Dólar hoje: {dolar}")

    r.close()
except Exception as e:
    print("Erro ao coletar cotação:", e)
    dolar = "Erro"

# === ETAPA 3: Gravar valor na planilha ===
data_hoje = dt.date.today().strftime('%d/%m/%y')
ws['A7'] = f"Cotação do dólar em {data_hoje}:"
ws['B7'] = dolar
ws['B7'].fill = PatternFill("solid", fgColor="00FFFF00")  # amarelo

# === ETAPA 4: Salvar arquivo Excel ===
wb.save(nome_arquivo)
print(f"Arquivo salvo em: {nome_arquivo}")
