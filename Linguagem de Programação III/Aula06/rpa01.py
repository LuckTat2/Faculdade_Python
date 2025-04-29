import rpa as r
import pyautogui as p
import pandas as pd
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# === CONFIGURAÇÕES DE CAMINHO ===
PASTA = r'C:\RPA'
CSV_TEMP = os.path.join(PASTA, 'cotacoes.csv')
CSV_FINAL = os.path.join(PASTA, 'bolsa.csv')
XLSX_FINAL = os.path.join(PASTA, 'bolsa.xlsx')

# Garante que a pasta existe
os.makedirs(PASTA, exist_ok=True)

# === ETAPA 1: ACESSA O SITE E FAZ LOGIN ===
r.init(visual_automation=True, chrome_browser=True)
r.url('https://br.advfn.com/ferramentas-de-investimento/monitor')
time.sleep(5)

janela = p.getActiveWindow()
janela.maximize()
time.sleep(2)

# Login
r.type('//*[@id="afnmainbodid"]/div[1]/div[2]/div[2]/div/form/div[2]/input', 'Teste')
time.sleep(2)
r.type('//*[@id="password-input"]', 'Teste1234')
time.sleep(1)
r.click('//*[@id="afnmainbodid"]/div[1]/div[2]/div[2]/div/form/input[1]')
time.sleep(5)

# Acessa o monitor com ações salvas
r.click('//*[@id="monitor-selector"]/div/ul/li[2]/button')
time.sleep(3)
p.click(x=76, y=287)  # Pode precisar ajustar essa coordenada!
time.sleep(6)

# === ETAPA 2: CAPTURA A TABELA E SALVA EM CSV ===
r.table('//*[@id="monitorApp_monGrid_grid_table"]', CSV_TEMP)
r.close()

# === ETAPA 3: CONVERTE PARA EXCEL COM FORMATAÇÃO ===
try:
    df = pd.read_csv(CSV_TEMP)
    df.to_csv(CSV_FINAL, index=False)  # Evita duplicações
    df.to_excel(XLSX_FINAL, index=False)

    # Formatação do Excel
    wb = load_workbook(XLSX_FINAL)
    ws = wb.active

    # Formata cabeçalhos
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True, color="000000")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajuste de largura automática das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(XLSX_FINAL)
    print(f"✅ Excel salvo em: {XLSX_FINAL}")

    # Abre o arquivo Excel
    os.startfile(XLSX_FINAL)

except Exception as e:
    print("❌ Erro ao processar Excel:", e)
