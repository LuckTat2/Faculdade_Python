import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Dados da planilha
dados = [
    ['Lucas Silva', 'lucas.desouza@aluno.fmpsc.edu.br', '48996411499', '2025-04-20 00:00:00', 150, 'Pendente'],
    ['Lucas Silva', 'lucas.desouza@aluno.fmpsc.edu.br', '48996411499', '2025-04-21 00:00:00', 200, 'Pago'],
    ['Lucas Silva', 'lucas.desouza@aluno.fmpsc.edu.br', '48996411499', '2025-04-20 00:00:00', 250, 'Pendente'],
    ['Lucas Silva', 'lucas.desouza@aluno.fmpsc.edu.br', '48996411499', '2025-04-20 00:00:00', 350, 'Pendente']
]

colunas = ['Nome', 'Email', 'Telefone', 'Vencimento', 'Valor', 'Status']

# Cria o DataFrame
df = pd.DataFrame(dados, columns=colunas)

# Caminho de saída
output_path = r'C:\RPA\AV01\clientes.xlsx'

# Exporta planilha inicial
df.to_excel(output_path, index=False)

# ---------- Estilização com openpyxl ----------
wb = load_workbook(output_path)
ws = wb.active

# Estilos
fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Aplica estilos
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        # Cabeçalho em negrito
        if cell.row == 1:
            cell.font = bold_font

        # Centraliza texto e aplica borda
        cell.alignment = center_align
        cell.border = thin_border

        # Formata a célula de status
        if cell.column == colunas.index('Status') + 1 and cell.row > 1:
            if cell.value == 'Pago':
                cell.fill = fill_verde
            elif cell.value == 'Pendente':
                cell.fill = fill_vermelho

# Salva a planilha formatada
wb.save(output_path)

print(f"Tabela formatada com sucesso em: {output_path}")
