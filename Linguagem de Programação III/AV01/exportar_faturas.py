import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Conexão com o banco de dados
user = 'root'
password = 'admin121'
host = 'localhost'
port = 3306
database = 'sistema_cobranca'

engine = create_engine(f'mysql+pymysql://{user}:{password}@{host}:{port}/{database}')

# Consulta SQL
query = """
SELECT f.id, c.nome, c.email, c.telefone, f.valor, f.vencimento, f.status
FROM faturas f
JOIN clientes c ON f.cliente_id = c.id
"""

# Carrega os dados
df = pd.read_sql(query, engine)

# Nova coluna "Status" legível
df['Status'] = df['status'].apply(lambda x: 'Pago' if x.lower() == 'pago' else 'Pendente')
df.drop(columns=['status'], inplace=True)

# Ordena por vencimento
df['vencimento'] = pd.to_datetime(df['vencimento'])
df.sort_values(by='vencimento', inplace=True)

# Caminho do arquivo
output_path = r'C:\RPA\AV01\faturas_exportadas.xlsx'

# Cria o ExcelWriter
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Aba Faturas
    df.to_excel(writer, sheet_name='Faturas', index=False)

    # Aba Resumo com cliente, status e valores
    resumo_completo = df.groupby(['nome', 'Status'])['valor'].sum().reset_index()
    resumo_completo.columns = ['Cliente', 'Status', 'Valor Total (R$)']
    resumo_completo.to_excel(writer, sheet_name='Resumo', index=False)

# Reabre para aplicar formatação
wb = load_workbook(output_path)
ws_faturas = wb['Faturas']
ws_resumo = wb['Resumo']

# Estilos
verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
bold = Font(bold=True)
center = Alignment(horizontal='center', vertical='center')
borda = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def aplicar_formatacao(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                cell.font = bold
            cell.alignment = center
            cell.border = borda

    col_status = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == 'Status':
            col_status = i
            break

    if col_status:
        for row in ws.iter_rows(min_row=2, min_col=col_status, max_col=col_status):
            for cell in row:
                if cell.value == 'Pago':
                    cell.fill = verde
                elif cell.value == 'Pendente':
                    cell.fill = vermelho

# Aplica nas duas abas
aplicar_formatacao(ws_faturas)
aplicar_formatacao(ws_resumo)

# Salva resultado final
wb.save(output_path)

print(f"Relatório exportado com sucesso para: {output_path}")
